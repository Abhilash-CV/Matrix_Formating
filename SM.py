import streamlit as st
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import io

st.set_page_config(layout="wide")

st.markdown("<h2 style='text-align:center;'>Category-wise distribution of the Seats in PHARMACY Colleges</h2>", unsafe_allow_html=True)
st.write("")

# ============================
# COLLEGE MAPPING (Your Mapping)
# ============================
college_map = {
    "ALB": "College of Pharmaceutical Sciences, Alappuzha",
    "AMB": "College of Pharmaceutical Sciences, Kannur",
    "KKB": "College of Pharmaceutical Sciences, Kozhikode",
    "KTB": "College of Pharmaceutical Sciences, Kottayam",
    "TVB": "College of Pharmaceutical Sciences, Thiruvananthapuram",
    "AAB": "Al-Azhar College of Pharmacy, Thodupuzha",
    "AFB": "Al Shifa College of Pharmacy, Perinthalmanna",
    "AHB": "Ahalia School of Pharmacy, Palakkad",
    "CAB": "Caritas College of Pharmacy, Kottayam",
    "CCB": "Chemists College of Pharmaceutical Sciences, Ernakulam",
    "CHB": "MGM College of Pharmaceutical Sciences, Malappuram",
    "CRB": "Crescent College of Pharmaceutical Sciences, Kannur",
    "DAB": "Devaki Amma Memorial College of Pharmacy, Malappuram",
    "DMB": "Dr. Moopen's College of Pharmacy, Wayanad",
    "DPB": "Department of Pharmaceutical Science, Ettumanoor",
    "DVB": "Dale View College of Pharmacy, TVM",
    "ECB": "Ezhuthachan College of Pharmaceutical Science, TVM",
    "EPB": "Elims College of Pharmacy, Thrissur",
    "GCB": "Grace College of Pharmacy, Palakkad",
    "HGB": "Holy Grace Academy of Pharmacy, Thrissur",
    "HKB": "Hindustan College of Pharmacy, Kottayam",
    "JCB": "St. James College of Pharmaceutical Sciences, Chalakudy",
    "JDB": "JDT Islam College of Pharmacy, Kozhikode",
    "JSB": "Jamia Salafiya Pharmacy College, Malappuram",
    "KAB": "Kerala Academy of Pharmacy, Thiruvananthapuram",
    "KCB": "KMCT College of Pharmacy, Malappuram",
    "KEB": "Indira Gandhi Institute of Pharmaceutical Sciences, Ernakulam",
    "KKP": "KMCT Institute of Pharmaceutical Education & Research, Kozhikode",
    "KLB": "KMCT Institute of Pharmacy, Malappuram",
    "KMB": "KMCT College of Pharmaceutical Sciences, Kozhikode",
    "KNB": "College of Pharmacy, Kannur Medical College",
    "KPB": "KTN College of Pharmacy, Palakkad",
    "KRB": "Karuna College of Pharmacy, Palakkad",
    "KVB": "KVM College of Pharmacy, Cherthala",
    "LIB": "Lisie College of Pharmacy, Ernakulam",
    "MAB": "Malabar College of Pharmacy, Malappuram",
    "MCB": "Moulana College of Pharmacy, Malappuram",
    "MDB": "Mar Dioscorus College of Pharmacy, TVM",
    "MEB": "MET's College of Pharmaceutical Sciences, Thrissur",
    "MGB": "MGM Silver Jubilee College of Pharmacy, Ernakulam",
    "MGK": "MGM Silver Jubilee College of Pharmacy, TVM",
    "MGR": "MGM College of Pharmacy, Kannur",
    "MKB": "Malik Deenar College of Pharmacy, Kasaragod",
    "MLB": "Madin College of Pharmacy, Malappuram",
    "MMB": "Mookambika College of Pharmaceutical Sciences, Muvattupuzha",
    "MPB": "Dr Joseph Mar Thoma Institute, Kattanam",
    "MZB": "Mount Zion College of Pharmacy, Adoor",
    "NCB": "Nehru College of Pharmacy, Thrissur",
    "NAB": "Nazareth College of Pharmacy, Pathanamthitta",
    "NEB": "Nirmala College of Health Science, Thrissur",
    "NMB": "Nirmala College of Pharmacy, Ernakulam",
    "NPB": "National College of Pharmacy, Kozhikode",
    "PCB": "Pushpagiri College of Pharmacy, Pathanamthitta",
    "PPB": "Prime College of Pharmacy, Palakkad",
    "RGB": "Rajiv Gandhi Institute of Pharmacy, Kasaragod",
    "RIB": "Department of Pharmaceutical Science, Kottayam",
    "SCB": "St Joseph's College of Pharmacy, Alappuzha",
    "SJB": "St John's College of Pharmaceutical Sciences, Idukki",
    "SKP": "Sree Krishna College of Pharmacy, TVM",
    "SPB": "Sanjoe College of Pharmaceutical Studies, Palakkad",
    "STB": "Sree Gokulam SNGM College of Pharmacy, Alappuzha",
    "TIB": "Triveni Institute of Pharmacy, Thrissur",
    "WSB": "Westfort College of Pharmacy, Thrissur"
}

uploaded = st.file_uploader("Upload Seat Matrix Excel", type=["xlsx"])

# ============================================================
# PROCESS INPUT
# ============================================================
if uploaded:

    df = pd.read_excel(uploaded)
    df = df[df["Program"].notna()]

    CATEGORY_MAP = {
        "SM":"SM","EW":"EWS","SC":"SC","ST":"ST",
        "EZ":"EZ","MU":"MU","BH":"BH","LA":"LA",
        "BX":"BX","KU":"KU","DV":"SQ","KN":"SQ"
    }

    MAIN_COLS = [
        "Course","Seats","SQ","SM","EWS","SC","ST","EZ","MU",
        "BH","LA","BX","KU",
        "SM-PD","EWS-PD","SC-PD","ST-PD","EZ-PD",
        "MU-PD","BH-PD","LA-PD","BX-PD"
    ]

    rows=[]
    for _,r in df.iterrows():
        row={"College":r["College"],"Course":r["Specialty"]}
        for old,new in CATEGORY_MAP.items():
            row[new]=r.get(old,0)
        for pd in ["SM-PD","EWS-PD","SC-PD","ST-PD","EZ-PD","MU-PD","BH-PD","LA-PD","BX-PD"]:
            row[pd]=0
        row["Seats"]=sum(row.get(c,0) for c in ["SQ","SM","EWS","SC","ST","EZ","MU","BH","LA","BX","KU"])
        rows.append(row)

    result=pd.DataFrame(rows)

    # ============================================================
    # DISPLAY IN STREAMLIT
    # ============================================================
    for col in result["College"].unique():
        cname = college_map.get(col, col)
        st.markdown(f"<h4 style='text-align:center;'>{col} : {cname}</h4>", unsafe_allow_html=True)
        block = result[result["College"]==col][MAIN_COLS]
        total = block.sum(numeric_only=True)
        total["Course"]="Total"
        block = pd.concat([block,total.to_frame().T],ignore_index=True)
        st.dataframe(block,use_container_width=True)

    # ============================================================
    # CREATE EXCEL WITH FORMATTING
    # ============================================================
    wb=openpyxl.Workbook()
    ws=wb.active
    ws.title="Seat Distribution"

    border=Border(left=Side(style='thin'),right=Side(style='thin'),
                  top=Side(style='thin'),bottom=Side(style='thin'))

    fill_college=PatternFill(fgColor="C65911",fill_type="solid")
    fill_header=PatternFill(fgColor="F4B183",fill_type="solid")
    fill_highlight=PatternFill(fgColor="F8CBAD",fill_type="solid")

    font_normal=Font(name="Times New Roman",size=12)
    font_bold=Font(name="Times New Roman",size=12,bold=True)
    font_white_bold=Font(name="Times New Roman",size=12,bold=True,color="FFFFFF")

    row_pos=1

    # ============================================================
    # PER COLLEGE BLOCK
    # ============================================================
    for col in result["College"].unique():

        cname = college_map.get(col, col)

        # COLLEGE TITLE
        ws.merge_cells(start_row=row_pos,start_column=1,end_row=row_pos,end_column=len(MAIN_COLS))
        t=ws.cell(row=row_pos,column=1)
        t.value=f"{col}: {cname}"
        t.font=font_white_bold
        t.fill=fill_college
        t.alignment=Alignment(horizontal="center")
        row_pos+=1

        # HEADER ROW
        for i,h in enumerate(MAIN_COLS,start=1):
            c=ws.cell(row=row_pos,column=i)
            c.value=h
            c.font=font_bold
            c.fill=fill_header
            c.border=border
            c.alignment=Alignment(horizontal="center")
        row_pos+=1

        # DATA ROWS
        block = result[result["College"]==col][MAIN_COLS]
        total = block.sum(numeric_only=True)
        total["Course"]="Total"
        block = pd.concat([block, total.to_frame().T], ignore_index=True)

        for r in block.itertuples(index=False):
            row_pos+=1
            for idx,v in enumerate(r,start=1):
                cell=ws.cell(row=row_pos,column=idx)
                cell.value=v
                cell.border=border
                cell.alignment=Alignment(horizontal="center")

                if isinstance(v,(int,float)) and v>0:
                    cell.fill=fill_highlight
                    cell.font=font_bold
                else:
                    cell.font=font_normal

        row_pos+=2

    # ============================================================
    # GRAND TOTAL SECTION
    # ============================================================
    grand = result[MAIN_COLS].sum(numeric_only=True)
    grand["Course"] = "Grand Total"
    grand_df = pd.DataFrame([grand])

    # GRAND TITLE
    ws.merge_cells(start_row=row_pos,start_column=1,end_row=row_pos,end_column=len(MAIN_COLS))
    gtitle = ws.cell(row=row_pos,column=1)
    gtitle.value = "GRAND TOTAL"
    gtitle.font = font_white_bold
    gtitle.fill = fill_college
    gtitle.alignment = Alignment(horizontal="center")
    row_pos += 1

    # GRAND HEADER
    for i, h in enumerate(MAIN_COLS, start=1):
        c = ws.cell(row=row_pos, column=i)
        c.value = h
        c.font = font_bold
        c.fill = fill_header
        c.border = border
        c.alignment = Alignment(horizontal="center")
    row_pos += 1

    # GRAND TOTAL ROW
    for r in grand_df.itertuples(index=False):
        row_pos += 1
        for idx, v in enumerate(r, start=1):
            cell = ws.cell(row=row_pos, column=idx)
            cell.value = v
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

            if isinstance(v, (int, float)) and v > 0:
                cell.fill = fill_highlight
                cell.font = font_bold
            else:
                cell.font = font_normal

    # ============================================================
    # DOWNLOAD BUTTON
    # ============================================================
    buf=io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    st.download_button(
        "📥 Download Excel (Formatted – Single Sheet + GRAND TOTAL)",
        data=buf.getvalue(),
        file_name="Seat_Distribution.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
